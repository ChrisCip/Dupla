"""Fusión de PDFs para visión y PDF de texto simple desde pliego Excel."""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Iterable

logger = logging.getLogger("dupla.nasas_pdf_utils")


def merge_pdfs(inputs: Iterable[Path], output: Path) -> Path:
    """Concatena PDFs en orden. Requiere PyMuPDF."""
    import fitz  # pymupdf

    paths = [Path(p).resolve() for p in inputs if Path(p).is_file()]
    if not paths:
        raise ValueError("merge_pdfs: no hay archivos válidos")
    output.parent.mkdir(parents=True, exist_ok=True)
    merged = fitz.open()
    try:
        for p in paths:
            with fitz.open(p) as doc:
                merged.insert_pdf(doc)
        merged.save(str(output))
    finally:
        merged.close()
    logger.info("PDF fusionado (%d archivos) → %s", len(paths), output)
    return output.resolve()


def pliego_xlsx_to_text_pdf(xlsx_path: Path, output_pdf: Path, *, max_chars: int = 45000) -> Path:
    """
    Exporta texto del primer libro (todas las hojas) a un PDF multipágina legible para contexto.
    No sustituye un pliego maquetado: sirve para que la visión GPT reciba condiciones en texto.
    """
    import fitz
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    lines: list[str] = []
    for name in wb.sheetnames:
        sh = wb[name]
        lines.append(f"=== Hoja: {name} ===")
        for row in sh.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not any(cells):
                continue
            line = " | ".join(c for c in cells if c)
            if line:
                lines.append(line)
    text = "\n".join(lines)
    text = text[:max_chars]
    output_pdf.parent.mkdir(parents=True, exist_ok=True)

    doc = fitz.open()
    try:
        margin = 56
        fontsize = 9
        page_w, page_h = fitz.paper_rect("a4")
        rect = fitz.Rect(margin, margin, page_w - margin, page_h - margin)
        remaining = text
        while remaining.strip():
            page = doc.new_page(width=page_w, height=page_h)
            block, remaining = _take_text_block(remaining, rect, fontsize)
            page.insert_textbox(
                rect,
                block,
                fontsize=fontsize,
                fontname="helv",
                align=fitz.TEXT_ALIGN_LEFT,
            )
        doc.save(str(output_pdf))
    finally:
        doc.close()
    return output_pdf.resolve()


def _take_text_block(text: str, rect, fontsize: float) -> tuple[str, str]:
    """Parte aproximada por longitud; el textbox de fitz hace el ajuste final."""
    approx = int(rect.width * rect.height / (fontsize * fontsize * 0.45))
    approx = max(800, approx)
    if len(text) <= approx:
        return text, ""
    # cortar en salto de línea cercano
    chunk = text[:approx]
    cut = chunk.rfind("\n")
    if cut < approx // 2:
        cut = approx
    return text[:cut].rstrip(), text[cut:].lstrip()


def collect_pdfs_recursive(root: Path, *, exclude_names: frozenset[str] | None = None) -> list[Path]:
    """Lista PDFs bajo root, ordenados por ruta (estable)."""
    ex = exclude_names or frozenset()
    out: list[Path] = []
    for p in sorted(root.rglob("*.pdf")):
        if p.name.startswith("~$"):
            continue
        if p.name in ex:
            continue
        out.append(p.resolve())
    return out


def sanitize_filename(name: str) -> str:
    s = re.sub(r'[<>:"/\\|?*]', "_", name)
    return s.strip()[:180] or "file"

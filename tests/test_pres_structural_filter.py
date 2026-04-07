from pathlib import Path

import pytest
from openpyxl import Workbook

from budget.pres_structural_filter import (
    filter_pres_workbook_structural,
    row_is_structural_chapter,
    row_is_structural_partida,
)


def test_row_is_structural_partida_positive() -> None:
    assert row_is_structural_partida("HZ01", "Partida", "Hormigón armado en zapatas fc=280 kg/cm2")


def test_row_is_structural_partida_negative_finish() -> None:
    assert not row_is_structural_partida("P05", "Partida", "Pañete en muros interiores e=1.75cm")


def test_row_is_structural_chapter() -> None:
    assert row_is_structural_chapter("Capítulo", "HORMIGÓN ARMADO")
    assert not row_is_structural_chapter("Capítulo", "INSTALACIONES ELÉCTRICAS")


def test_filter_pres_workbook_structural(tmp_path: Path) -> None:
    src = tmp_path / "pres.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Título"
    ws["A2"] = "Sub"
    ws["A3"] = "Headers"
    rows = [
        ("C1", "Capítulo", "ud", "HORMIGÓN ARMADO", "", "", ""),
        ("HZ1", "Partida", "m3", "Hormigón en zapatas", 10, 100, 1000),
        ("P05", "Partida", "m2", "Pañete interior", 5, 2, 10),
    ]
    for i, row in enumerate(rows, start=4):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    wb.save(src)

    dest = tmp_path / "out.xlsx"
    stats = filter_pres_workbook_structural(src, dest)
    assert stats["partidas_kept"] == 1
    assert dest.is_file()

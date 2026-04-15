from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

from budget.composer import compose_budget_rows
from budget.export_excel import export_budget_workbook
from core.schemas import BudgetCandidate, ProjectContext, QuantityTakeoff, QuantityTrace


def _sample_export_inputs() -> tuple[ProjectContext, list[QuantityTakeoff], dict[str, list[BudgetCandidate]]]:
    context = ProjectContext(project_id="demo_export", project_name="Demo Export Budget")
    takeoffs = [
        QuantityTakeoff(
            item_key="beam_01:concrete_volume",
            item_type="beam_concrete_volume",
            level_id="level_01",
            unit="m3",
            quantity=10.0,
            formula="beam.volume",
            inputs={
                "material_hint": "concrete",
                "context_tags": ["structural", "beam", "concrete", "volume"],
            },
            trace=QuantityTrace(
                source_entity_ids=["beam_01"],
                source_entity_sources=["hybrid"],
                metadata={
                    "material_hint": "concrete",
                    "context_tags": ["structural", "beam", "concrete", "volume"],
                },
            ),
        ),
        QuantityTakeoff(
            item_key="wall_01:paint",
            item_type="wall_finish_paint",
            level_id="level_01",
            unit="m2",
            quantity=18.0,
            formula="wall.finish",
            inputs={"context_tags": ["wall", "finish", "paint", "interior", "dry_area"]},
            trace=QuantityTrace(
                source_entity_ids=["wall_01"],
                source_entity_sources=["hybrid"],
                metadata={"context_tags": ["wall", "finish", "paint", "interior", "dry_area"]},
            ),
        ),
    ]
    candidates = {
        "beam_01:concrete_volume": [
            BudgetCandidate(
                takeoff_key="beam_01:concrete_volume",
                bc3_code="E300100",
                summary="Hormigon armado en vigas",
                unit="m3",
                score=0.81,
                rationale="Strong match.",
            )
        ],
        "wall_01:paint": [],
    }
    return context, takeoffs, candidates


def test_budget_export_excel_writes_workbook_with_headers_and_formulas() -> None:
    context, takeoffs, candidates = _sample_export_inputs()
    _, _, rows = compose_budget_rows(context, takeoffs, candidates)

    output_path = Path("tests") / "_budget_ready_test.xlsx"
    if output_path.exists():
        output_path.unlink()
    try:
        export_budget_workbook(context, rows, output_path)

        workbook = load_workbook(output_path, data_only=False)
        worksheet = workbook["Presupuesto"]

        assert worksheet["A1"].value == "Demo Export Budget"
        assert worksheet["A2"].value == "Presupuesto"
        assert [worksheet.cell(row=3, column=index).value for index in range(1, 8)] == [
            "Código",
            "Nat",
            "Ud",
            "Resumen",
            "CanPres",
            "PrPres",
            "ImpPres",
        ]

        chapter_row = None
        detail_row = None
        subtotal_row = None
        for row_index in range(4, worksheet.max_row + 1):
            nat = worksheet.cell(row=row_index, column=2).value
            if nat == "Capítulo" and chapter_row is None:
                chapter_row = row_index
            elif nat == "Partida" and detail_row is None:
                detail_row = row_index
            elif nat == "Subtotal/Cierre de capítulo" and subtotal_row is None:
                subtotal_row = row_index

        assert chapter_row is not None
        assert detail_row is not None
        assert subtotal_row is not None

        assert worksheet.cell(row=chapter_row, column=7).value.startswith("=G")
        assert worksheet.cell(row=detail_row, column=6).value is None
        assert worksheet.cell(row=detail_row, column=7).value.startswith("=ROUND(E")
        assert worksheet.cell(row=subtotal_row, column=6).value.startswith("=SUM(")
        assert worksheet.cell(row=subtotal_row, column=7).value.startswith("=ROUND(E")
        workbook.close()
    finally:
        if output_path.exists():
            output_path.unlink()


def test_budget_export_excel_falls_back_to_new_filename_when_target_is_locked(monkeypatch) -> None:
    context, takeoffs, candidates = _sample_export_inputs()
    _, _, rows = compose_budget_rows(context, takeoffs, candidates)

    requested_output = Path("tests") / "_budget_ready_locked_test.xlsx"
    original_save = OpenpyxlWorkbook.save
    save_calls: list[Path] = []

    if requested_output.exists():
        requested_output.unlink()
    for fallback_path in requested_output.parent.glob(f"{requested_output.stem}_*.xlsx"):
        fallback_path.unlink()

    def flaky_save(self, filename) -> None:
        path = Path(filename)
        save_calls.append(path)
        if path == requested_output:
            raise PermissionError("File is locked")
        return original_save(self, filename)

    monkeypatch.setattr(OpenpyxlWorkbook, "save", flaky_save)

    try:
        actual_output = export_budget_workbook(context, rows, requested_output)

        assert actual_output != requested_output
        assert actual_output.exists()
        assert actual_output.parent == requested_output.parent
        assert actual_output.suffix == ".xlsx"
        assert actual_output.stem.startswith(f"{requested_output.stem}_")
        assert save_calls[0] == requested_output
        assert save_calls[-1] == actual_output

        workbook = load_workbook(actual_output, data_only=False)
        try:
            assert workbook["Presupuesto"]["A1"].value == "Demo Export Budget"
        finally:
            workbook.close()
    finally:
        if requested_output.exists():
            requested_output.unlink()
        for fallback_path in requested_output.parent.glob(f"{requested_output.stem}_*.xlsx"):
            fallback_path.unlink()


def test_budget_export_excel_adds_quality_report_sheet_when_provided() -> None:
    context, takeoffs, candidates = _sample_export_inputs()
    _, _, rows = compose_budget_rows(context, takeoffs, candidates)
    output_path = Path("tests") / "_budget_quality_sheet.xlsx"
    if output_path.exists():
        output_path.unlink()

    quality_report = {
        "discipline": "arquitectura",
        "summary": {"total_elements": 2, "ok_count": 1, "warning_count": 0, "blocked_count": 1},
        "issues": [
            {
                "status": "BLOCKED",
                "code": "missing_space",
                "discipline": "arquitectura",
                "element_id": "wall_01",
                "level_id": "level_01",
                "unit_id": "level_01:unit_01",
                "space_id": None,
                "confidence_score": 0.4,
                "message": "Elemento sin espacio/unidad asignable con evidencia suficiente.",
                "evidence_refs": [],
                "suggested_action": "Agregar etiquetas espaciales en plano.",
            }
        ],
    }

    try:
        export_budget_workbook(context, rows, output_path, quality_report=quality_report)
        workbook = load_workbook(output_path, data_only=False)
        try:
            assert "Quality_Report" in workbook.sheetnames
            ws = workbook["Quality_Report"]
            assert ws["A1"].value == "status"
            assert ws["D2"].value == "wall_01"
            assert ws["I2"].value == "Elemento sin espacio/unidad asignable con evidencia suficiente."
        finally:
            workbook.close()
    finally:
        if output_path.exists():
            output_path.unlink()

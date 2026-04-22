from pathlib import Path

from openpyxl import Workbook

from analysis.day2_prep import build_day2_dataset_artifacts, select_validation_pairs
from knowledge.training_data import extract_training_pairs


def _write_sample_pres(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hoja1"
    sheet.append(["TORRE GIUALCA I"])
    sheet.append(["Presupuesto"])
    sheet.append(["Código", "Nat", "Ud", "Resumen", "CanPres", "PrPres", "ImpPres"])
    sheet.append(["TGIU0003", "Capítulo", "", "SEMISOTANO", 1, 1000, 1000])
    sheet.append(["TGIU00301", "Capítulo", "", "HORMIGON ARMADO", 1, 500, 500])
    sheet.append(["P0303130", "Partida", "m3", "Zapata Z1", 7.56, 399.4, 3019.46])
    sheet.append(["P030320S", "Partida", "m3", "Columna C1", 10, 1055.26, 10552.6])
    sheet.append(["TGIU00305", "Capítulo", "", "TERMINACIÓN DE SUPERFICIES", 1, 100, 100])
    sheet.append(["P0501101", "Partida", "m2", "Pañete en muros interiores", 653.87, 8.4, 5492.51])
    sheet.append(["TGIU0004", "Capítulo", "", "NIVEL 5", 1, 1200, 1200])
    sheet.append(["TGIU00406", "Capítulo", "", "TERMINACIÓN DE PISOS", 1, 300, 300])
    sheet.append(["P0610001", "Partida", "m2", "Piso Porcelanato Interior Apartamento", 20.37, 42.32, 862.06])
    workbook.save(path)


def _write_sample_pres_variant(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hoja1"
    sheet.append(["NASAS"])
    sheet.append(["Presupuesto"])
    sheet.append(["Código", "Nat", "Ud", "Resumen", "CanPres", "PrPres", "ImpPres"])
    sheet.append(["NAS0001", "Capítulo", "", "NIVEL 1", 1, 1000, 1000])
    sheet.append(["NAS0100", "Capítulo", "", "TERMINACIÓN DE SUPERFICIES", 1, 100, 100])
    sheet.append(["N010101", "Partida", "m2", "Pintura interior lavable", 10, 4.5, 45])
    workbook.save(path)


def _write_bad_format_pres(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hoja1"
    sheet.append(["BAD"])
    sheet.append(["Presupuesto"])
    sheet.append(["Código", "Nat", "Ud", "Resumen", "CanPres", "PrPres", "ImpPres"])
    sheet.append(["BAD0001", "Capítulo", "", "NIVEL 1", 1, 1000, 1000])
    sheet.append(["B01", "Partida", "Unidad demasiado larga para ser unidad", "Item con columnas corridas", 1, 1, 1])
    workbook.save(path)


def test_select_validation_pairs_balances_item_types(tmp_path: Path) -> None:
    pres_path = tmp_path / "PRES.xlsx"
    _write_sample_pres(pres_path)
    pairs = extract_training_pairs(pres_path)

    validation_pairs = select_validation_pairs(pairs, limit=3)

    assert len(validation_pairs) == 3
    assert len({pair.input_item_type for pair in validation_pairs}) >= 2


def test_build_day2_dataset_artifacts_writes_bundle(tmp_path: Path) -> None:
    pres_path = tmp_path / "PRES.xlsx"
    output_dir = tmp_path / "day2"
    _write_sample_pres(pres_path)

    manifest = build_day2_dataset_artifacts(pres_path, output_dir, validation_limit=2)

    assert manifest["train_records"] == 2
    assert manifest["validation_records"] == 2
    assert Path(manifest["train_path"]).exists()
    assert Path(manifest["validation_path"]).exists()
    assert Path(manifest["report_path"]).exists()

    report_text = Path(manifest["report_path"]).read_text(encoding="utf-8")
    assert "Day 2 training dataset report" in report_text
    assert "Validation examples" in report_text


def test_build_day2_dataset_artifacts_accepts_multiple_pres_sources(tmp_path: Path) -> None:
    pres_path_a = tmp_path / "PRES_A.xlsx"
    pres_path_b = tmp_path / "PRES_B.xlsx"
    output_dir = tmp_path / "day2_multi"
    _write_sample_pres(pres_path_a)
    _write_sample_pres_variant(pres_path_b)

    manifest = build_day2_dataset_artifacts(
        [pres_path_a, pres_path_b],
        output_dir,
        validation_limit=2,
    )

    assert len(manifest["pres_paths"]) == 2
    assert str(pres_path_a.resolve()) in manifest["source_pair_counts"]
    assert str(pres_path_b.resolve()) in manifest["source_pair_counts"]
    assert manifest["source_pair_counts"][str(pres_path_b.resolve())] == 1


def test_build_day2_dataset_artifacts_excludes_low_quality_source(tmp_path: Path) -> None:
    pres_good = tmp_path / "PRES_GOOD.xlsx"
    pres_bad = tmp_path / "PRES_BAD.xlsx"
    output_dir = tmp_path / "day2_filtered"
    _write_sample_pres(pres_good)
    _write_bad_format_pres(pres_bad)

    manifest = build_day2_dataset_artifacts(
        [pres_good, pres_bad],
        output_dir,
        validation_limit=2,
        min_source_quality=0.9,
    )

    assert str(pres_good.resolve()) in manifest["source_pair_counts"]
    assert str(pres_bad.resolve()) not in manifest["source_pair_counts"]
    assert str(pres_bad.resolve()) in manifest["excluded_sources"]
from io import BytesIO
from pathlib import Path
from uuid import UUID

from fpdf import FPDF
from fpdf.enums import XPos
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.models.user import User
from app.services.pliego_template_fill import (
    fill_pliego_workbook,
    resolve_pliego_template_path,
    suggested_pliego_xlsx_filename,
    workbook_to_bytes,
)
from app.services.project_service import ProjectService

settings = get_settings()


class ExportService:
    def __init__(self, session: AsyncSession) -> None:
        self._project_service = ProjectService(session)

    async def _load_payload(self, user: User, project_uuid: UUID) -> dict:
        payload, _ = await self._project_service.get_architecture(user, project_uuid)
        return payload

    def build_pliego_xlsx(self, payload: dict) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pliego"
        headers = ["Grupo", "Tipo", "Ítem", "Descripción", "Unidad", "Cant.", "P. Unit.", "Subtotal", "Notas"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
        row_idx = 2
        for g in payload.get("groups", []):
            gtitle = g.get("title", "")
            gkind = g.get("kind", "")
            for it in g.get("items", []):
                ws.cell(row=row_idx, column=1, value=gtitle)
                ws.cell(row=row_idx, column=2, value=gkind)
                ws.cell(row=row_idx, column=3, value=str(it.get("partida", "") or it.get("id", "")))
                ws.cell(row=row_idx, column=4, value=it.get("descripcion", ""))
                ws.cell(row=row_idx, column=5, value=it.get("unidad", ""))
                ws.cell(row=row_idx, column=6, value=it.get("cantidad"))
                ws.cell(row=row_idx, column=7, value=it.get("precio_unitario"))
                ws.cell(row=row_idx, column=8, value=it.get("subtotal"))
                ws.cell(row=row_idx, column=9, value=it.get("notas", ""))
                row_idx += 1
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def build_control_planos_xlsx(self, payload: dict) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Control Planos"
        headers = ["Grupo / Fase", "Plano / Referencia", "Descripción", "Estado"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = Font(bold=True)
        row_idx = 2
        for g in payload.get("groups", []):
            for it in g.get("items", []):
                ws.cell(row=row_idx, column=1, value=g.get("title", ""))
                ws.cell(row=row_idx, column=2, value=str(it.get("partida", "") or it.get("id", "")))
                ws.cell(row=row_idx, column=3, value=it.get("descripcion", ""))
                ws.cell(row=row_idx, column=4, value=it.get("notas", ""))
                row_idx += 1
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def build_pdf(self, title: str, payload: dict) -> bytes:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, title, ln=True)
        pdf.set_font("Helvetica", size=10)
        for g in payload.get("groups", []):
            pdf.set_font("Helvetica", "B", 11)
            pdf.multi_cell(0, 7, f"{g.get('title', '')} ({g.get('kind', '')})", new_x=XPos.LMARGIN)
            pdf.set_font("Helvetica", size=9)
            for it in g.get("items", []):
                line = (
                    f"{it.get('partida', '')} | {it.get('descripcion', '')} | "
                    f"{it.get('unidad', '')} | {it.get('cantidad', '')} | "
                    f"{it.get('precio_unitario', '')} | {it.get('subtotal', '')}"
                )
                pdf.multi_cell(0, 6, line, new_x=XPos.LMARGIN)
            pdf.ln(2)
        out = pdf.output()
        if isinstance(out, (bytes, bytearray)):
            return bytes(out)
        return str(out).encode("latin-1", errors="replace")

    async def export_pliego_xlsx(self, user: User, project_uuid: UUID) -> tuple[bytes, str]:
        payload = await self._load_payload(user, project_uuid)
        tpl = resolve_pliego_template_path(Path(settings.templates_dir))
        if tpl is not None:
            wb = load_workbook(tpl)
            if fill_pliego_workbook(wb, payload):
                return workbook_to_bytes(wb), suggested_pliego_xlsx_filename(str(project_uuid))
        return self.build_pliego_xlsx(payload), f"pliego-{project_uuid}.xlsx"

    async def export_control_xlsx(self, user: User, project_uuid: UUID) -> bytes:
        payload = await self._load_payload(user, project_uuid)
        tpl = Path(settings.templates_dir) / "GA-FO-03-control-planos.xlsx"
        if tpl.is_file():
            wb = load_workbook(tpl)
            return self._fill_template_control(wb, payload)
        return self.build_control_planos_xlsx(payload)

    def _fill_template_control(self, wb, payload: dict) -> bytes:
        _ = payload
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def export_pliego_pdf(self, user: User, project_uuid: UUID) -> bytes:
        payload = await self._load_payload(user, project_uuid)
        return self.build_pdf("Pliego de Condiciones - Arquitectura", payload)

    async def export_control_pdf(self, user: User, project_uuid: UUID) -> bytes:
        payload = await self._load_payload(user, project_uuid)
        return self.build_pdf("Control Entrega de Planos", payload)

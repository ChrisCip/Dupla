"""
PIPELINE INTEGRAL: DWG Analysis + Clash Detection + Budget por Partes.

Fases:
1. Leer archivos de aprendizaje (XLSX, BC3, RTF)
2. Analizar DWG nativo (COM) con propiedades completas
3. Clash Detection (nativo bounding-box + visual GPT-4o)
4. Presupuesto por partes (9 capitulos, 1 llamada GPT-4o cada uno)
5. Consolidar y generar salida TXT

Cada fase guarda data intermedia en dump/
"""

import win32com.client
import sys, os, json, re, time, base64
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

sys.path.insert(0, r"c:\Users\chris\Documents\Dupla")
from dotenv import load_dotenv
load_dotenv(r"c:\Users\chris\Documents\Dupla\.env")
from openai import OpenAI
from cad_automation.config import classify_layer

LEARNING_DIR = Path(r"c:\Users\chris\Documents\Dupla\learning_input")
OUTPUT_DIR = Path(r"c:\Users\chris\Documents\Dupla\analysis_output")
DUMP_DIR = OUTPUT_DIR / "dump"
PDF_DIR = Path(r"c:\Users\chris\Documents\Dupla\vision_output\pdf_pages")
DUMP_DIR.mkdir(parents=True, exist_ok=True)

ts = datetime.now().strftime("%Y%m%d_%H%M%S")
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

print("=" * 80)
print("  PIPELINE INTEGRAL DE ANALISIS DE OBRA")
print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 80)

# ============================================================
# FASE 1: LEER ARCHIVOS DE APRENDIZAJE
# ============================================================
print(f"\n{'='*70}\n[FASE 1/5] LEYENDO ARCHIVOS DE APRENDIZAJE\n{'='*70}")

# 1A: BC3
import openpyxl

def parse_bc3(filepath):
    raw = Path(filepath).read_text(encoding="latin-1", errors="replace")
    records = []
    cur = ""
    for line in raw.split("\n"):
        line = line.strip()
        if line.startswith("~"):
            if cur: records.append(cur)
            cur = line
        else:
            cur += line
    if cur: records.append(cur)
    
    concepts = {}
    hierarchy = defaultdict(list)
    texts = {}
    
    for r in records:
        if r.startswith("~C|"):
            parts = r[3:].split("|")
            if len(parts) < 2: continue
            code = parts[0].split("#")[0].strip()
            parents = parts[0].split("#",1)[1] if "#" in parts[0] else ""
            unit = parts[1].strip() if len(parts)>1 else ""
            summary = parts[2].strip() if len(parts)>2 else ""
            price = 0.0
            if len(parts)>3 and parts[3].strip():
                try: price = float(parts[3].strip())
                except: pass
            if code:
                concepts[code] = {"code":code,"unit":unit,"summary":summary,
                                  "price":price,"parents":parents}
        
        elif r.startswith("~D|"):
            parts = r[3:].split("|")
            if len(parts)<2: continue
            parent = parts[0].replace("#","").strip()
            for ch in parts[1].split("\\"):
                ch = ch.strip()
                if ch and ch != parent:
                    hierarchy[parent].append(ch)
        
        elif r.startswith("~T|"):
            parts = r[3:].split("|")
            if len(parts)>=2:
                code = parts[0].replace("#","").strip()
                if code: texts[code] = parts[1].strip()
    
    partidas = [c for c in concepts.values() if c["price"]>0 and c["unit"]]
    chapters = [c for c in concepts.values() if not c["unit"] and c["summary"]]
    return partidas, chapters, hierarchy, texts

bc3_files = list(LEARNING_DIR.glob("*.bc3")) + list(Path(r"c:\Users\chris\Documents\Dupla\presto_files").glob("*.bc3"))
all_partidas = []
all_chapters = []
all_hierarchy = {}

for f in bc3_files:
    print(f"  [BC3] {f.name}")
    p, c, h, t = parse_bc3(f)
    all_partidas.extend(p)
    all_chapters.extend(c)
    all_hierarchy.update(h)
    print(f"    {len(p)} partidas, {len(c)} capitulos")

# 1B: XLSX
def parse_xlsx(filepath):
    items = []
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        # Find header
        header_map = {}
        hrow = 0
        for i, row in enumerate(rows[:15]):
            txt = " ".join(str(c).lower() for c in row if c)
            if any(k in txt for k in ["partida","descripcion","unidad","cantidad","precio","total","codigo","concepto","importe"]):
                for j, cell in enumerate(row):
                    if cell is None: continue
                    ct = str(cell).lower().strip()
                    if any(k in ct for k in ["codigo","code","cod","nat"]): header_map["code"] = j
                    elif any(k in ct for k in ["descripcion","concepto","detalle","resumen"]): header_map["description"] = j
                    elif any(k in ct for k in ["unidad","ud"]): header_map["unit"] = j
                    elif any(k in ct for k in ["canpres","cantidad","cant"]): header_map["quantity"] = j
                    elif any(k in ct for k in ["precio","pres","p.u"]): header_map["unit_price"] = j
                    elif any(k in ct for k in ["imppres","total","importe","monto"]): header_map["total"] = j
                hrow = i
                break
        
        if not header_map:
            # Try using column index assuming standard layout
            for j, cell in enumerate(rows[0] if rows else []):
                if j == 0: header_map["code"] = j
                elif j == 1: header_map["description"] = j
                elif j == 2: header_map["unit"] = j
                elif j == 3: header_map["quantity"] = j
                elif j == 4: header_map["unit_price"] = j
                elif j == 5: header_map["total"] = j
        
        for row in rows[hrow+1:]:
            if not row or all(c is None for c in row): continue
            item = {"source": filepath.name, "sheet": sn}
            for field, ci in header_map.items():
                if ci < len(row) and row[ci] is not None:
                    if field in ("quantity","unit_price","total"):
                        try: item[field] = float(row[ci])
                        except: item[field] = 0.0
                    else:
                        item[field] = str(row[ci]).strip()
            if item.get("description") and len(item.get("description","")) > 2:
                items.append(item)
    wb.close()
    return items

xls_files = list(LEARNING_DIR.glob("*.xlsx")) + list(LEARNING_DIR.glob("*.xls"))
all_xls = []
for f in xls_files:
    print(f"  [XLS] {f.name}")
    items = parse_xlsx(f)
    all_xls.extend(items)
    print(f"    {len(items)} items")

# 1C: RTF files
rtf_texts = {}
for f in LEARNING_DIR.glob("*.RTF"):
    try:
        raw = f.read_text(encoding="latin-1", errors="replace")
        # Strip RTF formatting, keep text
        text = re.sub(r'\\[a-z]+[\d]*\s?', ' ', raw)
        text = re.sub(r'[{}]', '', text)
        text = re.sub(r'\s+', ' ', text).strip()
        rtf_texts[f.stem] = text[:500]
        print(f"  [RTF] {f.name}: {len(text)} chars")
    except:
        pass

# Save all learning data to dump
json.dump(all_partidas, open(DUMP_DIR/"partidas_bc3.json","w",encoding="utf-8"),
          ensure_ascii=False, indent=2)
json.dump(all_xls, open(DUMP_DIR/"items_xlsx.json","w",encoding="utf-8"),
          ensure_ascii=False, indent=2, default=str)
json.dump(all_chapters, open(DUMP_DIR/"chapters_bc3.json","w",encoding="utf-8"),
          ensure_ascii=False, indent=2)

print(f"\n  TOTAL: {len(all_partidas)} partidas BC3, {len(all_xls)} items XLS, {len(rtf_texts)} RTFs")

# ============================================================
# FASE 2: ANALISIS DWG NATIVO (COM)
# ============================================================
print(f"\n{'='*70}\n[FASE 2/5] ANALISIS DWG NATIVO (COM)\n{'='*70}")

acad = win32com.client.GetActiveObject("AutoCAD.Application")
doc = acad.ActiveDocument
msp = doc.ModelSpace
total = msp.Count

print(f"  Documento: {doc.Name}")
print(f"  Entidades: {total}")

layers = defaultdict(lambda: {
    "count":0, "types":Counter(), "total_length":0.0, "total_area":0.0,
    "min_x":float("inf"), "min_y":float("inf"),
    "max_x":float("-inf"), "max_y":float("-inf"),
    "bboxes": [],
})

start_t = time.time()
for i in range(total):
    if i % 5000 == 0 and i > 0:
        el = time.time()-start_t
        print(f"    [{i/total*100:.0f}%] {i}/{total} | {el:.0f}s")
    try:
        ent = msp.Item(i)
        ln = ent.Layer
        on = ent.ObjectName.replace("AcDb","")
        lay = layers[ln]
        lay["count"] += 1
        lay["types"][on] += 1
        
        try:
            if hasattr(ent,"Length") and on not in ("BlockReference","Text","MText"):
                try:
                    l = float(ent.Length)
                    if 0 < l < 100000: lay["total_length"] += l
                except: pass
            if hasattr(ent,"Area"):
                try:
                    a = float(ent.Area)
                    if 0 < a < 1000000: lay["total_area"] += a
                except: pass
            try:
                mn, mx = ent.GetBoundingBox()
                lay["min_x"] = min(lay["min_x"], mn[0])
                lay["min_y"] = min(lay["min_y"], mn[1])
                lay["max_x"] = max(lay["max_x"], mx[0])
                lay["max_y"] = max(lay["max_y"], mx[1])
                disc = classify_layer(ln).name
                lay["bboxes"].append({"disc": disc, "min": list(mn[:2]), "max": list(mx[:2])})
            except: pass
        except: pass
    except: pass

dwg_time = time.time()-start_t
print(f"  Completado en {dwg_time:.0f}s")

# Build dwg summary
dwg_layers = {}
for name, data in layers.items():
    w = data["max_x"]-data["min_x"] if data["max_x"]>data["min_x"] else 0
    h = data["max_y"]-data["min_y"] if data["max_y"]>data["min_y"] else 0
    disc = classify_layer(name).name
    dwg_layers[name] = {
        "name": name, "discipline": disc, "count": data["count"],
        "types": dict(data["types"]),
        "length_m": round(data["total_length"],2),
        "area_m2": round(data["total_area"],2),
        "bbox_w": round(w,2), "bbox_h": round(h,2),
    }

# Save DWG analysis
json.dump(dwg_layers, open(DUMP_DIR/"dwg_layers.json","w",encoding="utf-8"),
          ensure_ascii=False, indent=2)

# ============================================================
# FASE 3: CLASH DETECTION
# ============================================================
print(f"\n{'='*70}\n[FASE 3/5] CLASH DETECTION\n{'='*70}")

# 3A: Native clash detection (bounding box intersections)
print("  [3A] Deteccion nativa por bounding box...")

disc_bboxes = defaultdict(list)
for name, data in layers.items():
    disc = classify_layer(name).name
    for bb in data["bboxes"][:500]:  # Limit per layer
        disc_bboxes[disc].append({"layer": name, "min": bb["min"], "max": bb["max"]})

def bb_intersects(a, b):
    return (a["min"][0] <= b["max"][0] and a["max"][0] >= b["min"][0] and
            a["min"][1] <= b["max"][1] and a["max"][1] >= b["min"][1])

def overlap_pct(a, b):
    ix = max(0, min(a["max"][0],b["max"][0]) - max(a["min"][0],b["min"][0]))
    iy = max(0, min(a["max"][1],b["max"][1]) - max(a["min"][1],b["min"][1]))
    ia = ix * iy
    aw = (a["max"][0]-a["min"][0])*(a["max"][1]-a["min"][1])
    bw = (b["max"][0]-b["min"][0])*(b["max"][1]-b["min"][1])
    smaller = min(aw, bw) if min(aw,bw) > 0 else 1
    return ia / smaller * 100

clash_pairs = [("A","S"),("A","E"),("A","P"),("S","E"),("S","P"),("E","P")]
native_clashes = []

for d1, d2 in clash_pairs:
    bbs1 = disc_bboxes.get(d1, [])[:200]
    bbs2 = disc_bboxes.get(d2, [])[:200]
    count = 0
    for a in bbs1:
        for b in bbs2:
            if bb_intersects(a, b):
                ovlp = overlap_pct(a, b)
                if ovlp > 10:
                    count += 1
                    if len(native_clashes) < 500:
                        sev = "CRITICO" if ovlp > 80 else "MAYOR" if ovlp > 50 else "MENOR"
                        native_clashes.append({
                            "disc_a": d1, "layer_a": a["layer"],
                            "disc_b": d2, "layer_b": b["layer"],
                            "coord": [round((a["min"][0]+a["max"][0])/2,1),
                                      round((a["min"][1]+a["max"][1])/2,1)],
                            "overlap_pct": round(ovlp,1), "severity": sev,
                        })
    if count > 0:
        print(f"    {d1} vs {d2}: {count} clashes")

sev_count = Counter(c["severity"] for c in native_clashes)
print(f"  Total clashes nativos: {len(native_clashes)}")
print(f"  CRITICO:{sev_count.get('CRITICO',0)} MAYOR:{sev_count.get('MAYOR',0)} MENOR:{sev_count.get('MENOR',0)}")

json.dump(native_clashes, open(DUMP_DIR/"native_clashes.json","w",encoding="utf-8"),
          ensure_ascii=False, indent=2)

# 3B: Visual clash detection with GPT-4o
print("\n  [3B] Deteccion visual con GPT-4o...")
visual_clashes = []

page_images = sorted(PDF_DIR.glob("page_*.png"))[:5] if PDF_DIR.exists() else []

if page_images:
    for img_path in page_images:
        print(f"    Analizando {img_path.name}...")
        with open(img_path, "rb") as f:
            img_b64 = base64.b64encode(f.read()).decode()
        
        try:
            resp = client.chat.completions.create(
                model="gpt-4o", max_tokens=2048, temperature=0.1,
                messages=[
                    {"role":"system","content":"Ingeniero de coordinacion BIM. Identifica clashes visuales en planos."},
                    {"role":"user","content":[
                        {"type":"text","text":"Analiza este plano e identifica clashes (interferencias) entre disciplinas. Responde en JSON: {\"clashes\":[{\"severity\":\"CRITICO/MAYOR/MENOR\",\"disciplines\":\"A,S\",\"description\":\"que conflicto\",\"location\":\"donde\",\"recommendation\":\"solucion\"}]}"},
                        {"type":"image_url","image_url":{"url":f"data:image/png;base64,{img_b64}","detail":"low"}}
                    ]}
                ],
            )
            raw = resp.choices[0].message.content
            cleaned = re.sub(r'(?<=\d),(?=\d{3})', '', raw)
            m = re.search(r'\{.*\}', cleaned, re.DOTALL)
            if m:
                fixed = re.sub(r',\s*[}\]]', lambda x: x.group().replace(',',''), m.group())
                try:
                    result = json.loads(fixed)
                    for c in result.get("clashes", []):
                        c["source_page"] = img_path.name
                        visual_clashes.append(c)
                    print(f"      {len(result.get('clashes',[]))} clashes")
                except: pass
            time.sleep(0.5)
        except Exception as e:
            print(f"      [ERROR] {e}")
else:
    print("    No hay imagenes de paginas en pdf_pages/")

json.dump(visual_clashes, open(DUMP_DIR/"visual_clashes.json","w",encoding="utf-8"),
          ensure_ascii=False, indent=2)
print(f"  Total clashes visuales: {len(visual_clashes)}")

# ============================================================
# FASE 4: PRESUPUESTO POR PARTES
# ============================================================
print(f"\n{'='*70}\n[FASE 4/5] PRESUPUESTO POR PARTES (GPT-4o)\n{'='*70}")

# Build reference strings
dwg_summary = "MEDICIONES DWG:\n"
for name, d in sorted(dwg_layers.items(), key=lambda x: -x[1]["count"]):
    if d["count"] < 5: continue
    dwg_summary += f"  {name:<22} {d['discipline']:<6} {d['count']:>6} ent  L={d['length_m']:.1f}m  A={d['area_m2']:.1f}m2\n"

bc3_text = ""
for p in all_partidas[:200]:
    bc3_text += f"{p['code']:<16} {p['unit']:<6} {p['price']:>12,.2f}  {p['summary'][:55]}\n"

xls_text = ""
for item in all_xls[:100]:
    c = item.get("code","")
    d = item.get("description","")[:40]
    u = item.get("unit","")
    q = item.get("quantity",0)
    p = item.get("unit_price",0)
    t = item.get("total",0)
    if d:
        xls_text += f"  {c:<16} {d:<40} {u:<5} {q:>8.1f} {p:>10.1f} {t:>12.1f}\n"

def clean_and_parse(raw):
    cleaned = re.sub(r'(?<=\d),(?=\d{3})', '', raw)
    m = re.search(r"```(?:json)?\s*\n(.*?)\n```", cleaned, re.DOTALL)
    if m:
        try: return json.loads(m.group(1))
        except: pass
    try: return json.loads(cleaned)
    except: pass
    s, e = cleaned.find("{"), cleaned.rfind("}")
    if s >= 0 and e > s:
        fixed = re.sub(r',\s*[}\]]', lambda x: x.group().replace(',',''), cleaned[s:e+1])
        try: return json.loads(fixed)
        except: pass
    return None

categories = [
    ("01_MOVIMIENTO_TIERRAS", "Excavacion, relleno, compactacion, zapatas, bote de material, cimentacion"),
    ("02_ESTRUCTURA", "Hormigon armado, acero de refuerzo, columnas, vigas, losas, encofrado, escaleras"),
    ("03_MUROS_BLOQUES_PANETE", "Muros de bloques, panete interior y exterior, revestimientos ceramicos"),
    ("04_PISOS_CERAMICA", "Pisos ceramica, porcelanato, zocalos, pulido, nivelacion"),
    ("05_PUERTAS_VENTANAS", "Puertas metalicas, PVC, madera, ventanas aluminio, vidrios, herrajes, cerraduras"),
    ("06_ELECTRICAS", "Puntos electricos, cableado, interruptores, tomas, paneles, transformador, luminarias"),
    ("07_SANITARIAS_PLOMERIA", "Inodoros, lavamanos, duchas, tuberias PVC, cisterna, bombas, drenaje"),
    ("08_PINTURA_ACABADOS", "Pintura interior/exterior, impermeabilizante, sellador, acabados"),
    ("09_GASTOS_INDIRECTOS", "Supervision, topografia, seguridad, limpieza, andamios, desperdicios, equipos"),
]

budget_chapters = []
total_tokens = 0

for cat_code, cat_desc in categories:
    print(f"  [{cat_code}] {cat_desc[:50]}...")
    
    prompt = f"""Presupuestista senior dominicano. Capitulo: {cat_desc}

=== CATALOGO PRESTO BC3 (precios en RD$) ===
{bc3_text[:5000]}

=== PRESUPUESTO REFERENCIA XLS ===
{xls_text[:3000]}

=== MEDICIONES CAD (DWG) ===
{dwg_summary[:2500]}

Area total: 3,426 m2. Edificio 14 niveles + PH + semi-sotano.

REGLAS:
1. Usa codigos y precios del catalogo Presto/XLS
2. Cantidades del CAD cuando esten disponibles
3. Estima con area (3,426m2) y niveles (14) cuando no haya medicion directa
4. Incluye calculo de dimensiones para CADA partida
5. Sigue la estructura de capitulos del XLSX de referencia
6. SOLO partidas de: {cat_desc}

JSON (sin texto adicional):
{{"category":"{cat_code}","items":[{{"presto_code":"","description":"","unit":"","quantity":0,"unit_price":0,"total":0,"cad_layer":"","calculation":"formula o explicacion","match_type":"exacto/estimado"}}],"subtotal":0}}"""

    try:
        resp = client.chat.completions.create(
            model="gpt-4o", max_tokens=4096, temperature=0.1,
            messages=[
                {"role":"system","content":"Presupuestista dominicano. Solo JSON. Precios en RD$. Sigue estructura del XLS referencia."},
                {"role":"user","content":prompt},
            ],
        )
        raw = resp.choices[0].message.content
        tokens = resp.usage.total_tokens
        total_tokens += tokens
        
        # Save raw to dump
        (DUMP_DIR / f"raw_{cat_code}.txt").write_text(raw, encoding="utf-8")
        
        result = clean_and_parse(raw)
        if result:
            items = result.get("items",[])
            subtotal = sum(float(i.get("total",0)) for i in items)
            result["subtotal"] = subtotal
            budget_chapters.append(result)
            print(f"    {len(items)} partidas -> RD$ {subtotal:,.0f}")
        else:
            print(f"    [WARN] JSON parse failed")
        
        time.sleep(1)
    except Exception as e:
        print(f"    [ERROR] {e}")

# Save budget chapters
json.dump(budget_chapters, open(DUMP_DIR/"budget_chapters.json","w",encoding="utf-8"),
          ensure_ascii=False, indent=2, default=str)

# ============================================================
# FASE 5: GENERAR REPORTE FINAL
# ============================================================
print(f"\n{'='*70}\n[FASE 5/5] GENERANDO REPORTE FINAL\n{'='*70}")

grand_total = sum(ch.get("subtotal",0) for ch in budget_chapters)
total_items = sum(len(ch.get("items",[])) for ch in budget_chapters)
area = 3426
costo_m2 = grand_total / area if area > 0 else 0

lines = []

# Header
lines.append("=" * 115)
lines.append(f"  ANALISIS INTEGRAL DE OBRA Y PRESUPUESTO")
lines.append(f"  Proyecto: {doc.Name}")
lines.append(f"  Ubicacion: Santo Domingo, Republica Dominicana")
lines.append(f"  Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
lines.append(f"  Fuentes: DWG (COM) + BC3 (Presto) + XLSX (Referencia) + GPT-4o")
lines.append("=" * 115)

# PART 1: DWG Analysis
lines.append(f"\n{'#'*115}")
lines.append(f"  PARTE 1: ANALISIS DEL DWG")
lines.append(f"{'#'*115}")
lines.append(f"  Archivo:   {doc.Name}")
lines.append(f"  Entidades: {total:,}")
lines.append(f"  Tiempo:    {dwg_time:.0f}s")

disc_sum = defaultdict(lambda: {"count":0,"length":0,"area":0,"layers":[]})
for name, d in dwg_layers.items():
    ds = disc_sum[d["discipline"]]
    ds["count"] += d["count"]
    ds["length"] += d["length_m"]
    ds["area"] += d["area_m2"]
    ds["layers"].append(name)

lines.append(f"\n  {'Disc':<10} {'Ent':>8} {'Long (m)':>12} {'Area (m2)':>12}")
lines.append(f"  {'-'*10} {'-'*8} {'-'*12} {'-'*12}")
for d, s in sorted(disc_sum.items(), key=lambda x: -x[1]["count"]):
    lines.append(f"  {d:<10} {s['count']:>8,} {s['length']:>12,.1f} {s['area']:>12,.1f}")

lines.append(f"\n  DETALLE POR CAPA:")
lines.append(f"  {'Capa':<22} {'Disc':<6} {'Ent':>7} {'Long (m)':>12} {'Area (m2)':>12} {'Tipos'}")
lines.append(f"  {'-'*22} {'-'*6} {'-'*7} {'-'*12} {'-'*12} {'-'*30}")
for name, d in sorted(dwg_layers.items(), key=lambda x:-x[1]["count"]):
    if d["count"] < 5: continue
    tt = ", ".join(f"{k}:{v}" for k,v in sorted(d["types"].items(),key=lambda x:-x[1])[:3])
    lines.append(f"  {name:<22} {d['discipline']:<6} {d['count']:>7} "
                 f"{d['length_m']:>12,.1f} {d['area_m2']:>12,.1f} {tt}")

# PART 2: Clash Detection
lines.append(f"\n{'#'*115}")
lines.append(f"  PARTE 2: CLASH DETECTION")
lines.append(f"{'#'*115}")
lines.append(f"  Clashes nativos (bbox): {len(native_clashes)}")
lines.append(f"  Clashes visuales (IA):  {len(visual_clashes)}")
lines.append(f"  CRITICO: {sev_count.get('CRITICO',0)} | MAYOR: {sev_count.get('MAYOR',0)} | MENOR: {sev_count.get('MENOR',0)}")

if native_clashes:
    pair_counts = Counter(f"{c['disc_a']} vs {c['disc_b']}" for c in native_clashes)
    lines.append(f"\n  Por par de disciplinas:")
    for pair, cnt in pair_counts.most_common():
        lines.append(f"    {pair}: {cnt} clashes")
    
    lines.append(f"\n  Top 20 clashes criticos:")
    lines.append(f"  {'#':>4} {'Sev':<9} {'Disc':>5} {'Capa A':<20} {'Capa B':<20} {'Coord':<20} {'Overlap':>8}")
    for i, c in enumerate(native_clashes[:20]):
        lines.append(f"  {i+1:>4} {c['severity']:<9} {c['disc_a']}v{c['disc_b']:<3} "
                     f"{c['layer_a']:<20} {c['layer_b']:<20} "
                     f"({c['coord'][0]:.0f},{c['coord'][1]:.0f}){'':<8} {c['overlap_pct']:>6.1f}%")

if visual_clashes:
    lines.append(f"\n  Clashes visuales (GPT-4o):")
    for i, c in enumerate(visual_clashes):
        lines.append(f"  [{c.get('severity','?')}] {c.get('description','')}")
        lines.append(f"    Ubicacion: {c.get('location','')} | Pag: {c.get('source_page','')}")
        lines.append(f"    Recomendacion: {c.get('recommendation','')}")

# PART 3: Budget with all partidas
lines.append(f"\n{'#'*115}")
lines.append(f"  PARTE 3: PRESUPUESTO DETALLADO POR PARTIDAS")
lines.append(f"{'#'*115}")

for ch in budget_chapters:
    cat = ch.get("category","?")
    items = ch.get("items",[])
    subtotal = ch.get("subtotal",0)
    
    lines.append(f"\n{'_'*115}")
    lines.append(f"  {cat.replace('_',' ')}")
    lines.append(f"  Subtotal: RD$ {subtotal:>18,.2f}")
    lines.append(f"{'_'*115}")
    lines.append(f"  {'Codigo':<14} {'Descripcion':<38} {'Ud':<5} "
                 f"{'Cant':>10} {'P.Unit RD$':>14} {'Total RD$':>16} {'Tipo'}")
    lines.append(f"  {'-'*14} {'-'*38} {'-'*5} {'-'*10} {'-'*14} {'-'*16} {'-'*10}")
    
    for item in items:
        desc = str(item.get("description",""))[:36]
        tot = float(item.get("total",0))
        qty = float(item.get("quantity",0))
        prc = float(item.get("unit_price",0))
        lines.append(f"  {str(item.get('presto_code','')):<14} {desc:<38} "
                     f"{str(item.get('unit','')):<5} {qty:>10.2f} "
                     f"{prc:>14,.2f} {tot:>16,.2f} {str(item.get('match_type',''))[:10]}")
        calc = str(item.get("calculation",""))
        if calc:
            lines.append(f"{'':>16} CALC: {calc[:95]}")

# Summary
lines.append(f"\n{'='*115}")
lines.append(f"  RESUMEN GENERAL")
lines.append(f"{'='*115}")
for ch in budget_chapters:
    n = len(ch.get("items",[]))
    lines.append(f"  {ch.get('category','').replace('_',' '):<55} {n:>3} items  RD$ {ch.get('subtotal',0):>18,.2f}")
lines.append(f"  {'─'*55} {'─'*3}{'─'*5}  {'─'*22}")
lines.append(f"  {'TOTAL GENERAL':<55} {total_items:>3} items  RD$ {grand_total:>18,.2f}")
lines.append(f"\n  Area: {area:,} m2  |  Costo/m2: RD$ {costo_m2:,.2f}  |  Tokens: {total_tokens:,}")
lines.append(f"{'='*115}")

report = "\n".join(lines)
report_path = OUTPUT_DIR / f"ANALISIS_INTEGRAL_{ts}.txt"
report_path.write_text(report, encoding="utf-8")

# DWG-only report
dwg_report = "\n".join(lines[:lines.index(f"{'#'*115}", 2)])
dwg_path = OUTPUT_DIR / f"DWG_ANALYSIS_{ts}.txt"

# JSON
full_json = {
    "project": doc.Name, "date": ts,
    "dwg": dwg_layers, "dwg_entities": total,
    "native_clashes": native_clashes[:100],
    "visual_clashes": visual_clashes,
    "budget_chapters": budget_chapters,
    "grand_total": grand_total, "total_items": total_items,
    "costo_m2": costo_m2,
}
json_path = OUTPUT_DIR / f"ANALISIS_INTEGRAL_{ts}.json"
json.dump(full_json, open(json_path,"w",encoding="utf-8"),
          ensure_ascii=False, indent=2, default=str)

print(f"\n{'='*70}")
print(f"  PIPELINE INTEGRAL COMPLETADO!")
print(f"  Reporte:   {report_path.name}")
print(f"  JSON:      {json_path.name}")
print(f"  Capitulos: {len(budget_chapters)}")
print(f"  Partidas:  {total_items}")
print(f"  Clashes:   {len(native_clashes)} nativos + {len(visual_clashes)} visuales")
print(f"  Total:     RD$ {grand_total:,.2f}")
print(f"  Costo/m2:  RD$ {costo_m2:,.2f}")
print(f"{'='*70}")

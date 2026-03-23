"""
Excel export for composed Dupla budget rows.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.schemas import BudgetRow, ProjectContext

HEADERS = ("Código", "Nat", "Ud", "Resumen", "CanPres", "PrPres", "ImpPres")
THIN_SIDE = Side(style="thin", color="BFBFBF")
ALL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
CHAPTER_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")


def _coerce_row(row: BudgetRow | Mapping[str, object]) -> BudgetRow:
    if isinstance(row, BudgetRow):
        return row
    payload = dict(row)
    return BudgetRow(
        row_type=str(payload.get("row_type", "line")),
        code=str(payload.get("code", "")),
        nat=str(payload.get("nat", "")),
        unit=str(payload.get("unit", "")),
        summary=str(payload.get("summary", "")),
        quantity=payload.get("quantity"),
        unit_price=payload.get("unit_price"),
        amount=payload.get("amount"),
        chapter_id=payload.get("chapter_id"),
        parent_chapter_id=payload.get("parent_chapter_id"),
        level=int(payload.get("level", 0) or 0),
        takeoff_key=payload.get("takeoff_key"),
        source_refs=list(payload.get("source_refs", [])),
        assumptions=list(payload.get("assumptions", [])),
        metadata=dict(payload.get("metadata", {})),
        excel_row=payload.get("excel_row"),
    )


def _write_value(cell, value: object) -> None:
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        return
    if value is None:
        cell.value = None


def export_budget_workbook(
    context: ProjectContext,
    rows: Iterable[BudgetRow | Mapping[str, object]],
    output_path: str | Path,
    *,
    sheet_name: str = "Presupuesto",
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    project_title = context.project_name or context.project_id or "DUPLA"
    worksheet["A1"] = project_title
    worksheet["A2"] = "Presupuesto"

    worksheet["A1"].font = Font(size=14, bold=True)
    worksheet["A2"].font = Font(size=12, bold=True)

    for column_index, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(row=3, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = ALL_BORDER

    coerced_rows = [_coerce_row(row) for row in rows]
    for row in coerced_rows:
        target_row = row.excel_row or 4
        values = (
            row.code,
            row.nat,
            row.unit,
            row.summary,
            row.quantity,
            row.unit_price,
            row.amount,
        )
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=target_row, column=column_index)
            _write_value(cell, value)
            cell.border = ALL_BORDER
            if column_index >= 5:
                cell.number_format = '#,##0.00'

        row_fill = None
        row_font = Font(bold=False)
        if row.row_type == "chapter":
            row_fill = CHAPTER_FILL
            row_font = Font(bold=True)
        elif row.row_type == "subtotal":
            row_fill = SUBTOTAL_FILL
            row_font = Font(bold=True)

        for column_index in range(1, 8):
            cell = worksheet.cell(row=target_row, column=column_index)
            cell.font = row_font
            cell.alignment = Alignment(
                horizontal="left" if column_index <= 4 else "right",
                vertical="center",
            )
            if row_fill is not None:
                cell.fill = row_fill

    worksheet.freeze_panes = "A4"
    worksheet.sheet_view.showGridLines = True
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 16
    worksheet.column_dimensions["C"].width = 10
    worksheet.column_dimensions["D"].width = 60
    worksheet.column_dimensions["E"].width = 14
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 16

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output

"""
Lectura del formato «Preliminary Budget» (NASAS y similares): varias hojas,
cabecera con fila 'Código', columnas Resumen bilingüe y última columna de importe.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        cleaned = str(value).replace(",", ".").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0


def _find_header_row(sheet) -> int | None:
    for r in range(1, 40):
        v = sheet.cell(r, 1).value
        if v is not None and str(v).strip().lower() in {"código", "codigo"}:
            return r
    return None


def load_nasas_preliminary_budget_rows(path: Path) -> list[dict[str, Any]]:
    """
    Devuelve filas compatibles con compare_budget._load_budget_rows:
    code, nat, unit, summary, quantity, price, amount.
    """
    workbook = load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        hr = _find_header_row(sheet)
        if hr is None:
            continue
        for row in sheet.iter_rows(min_row=hr + 1, min_col=1, max_col=8, values_only=True):
            cells = list(row) + [None] * (8 - len(row))
            code = _safe_str(cells[0])
            nat = _safe_str(cells[1])
            summary_es = _safe_str(cells[2])
            summary_en = _safe_str(cells[3])
            qty = _safe_float(cells[4])
            unit = _safe_str(cells[5])
            col_g = cells[6]
            col_h = cells[7]
            price = _safe_float(col_g)
            amount = _safe_float(col_h) if col_h not in (None, "") else price

            if not any((code, nat, summary_es, summary_en)):
                continue
            summary = summary_es or summary_en
            if summary_en and summary_es and summary_en != summary_es:
                summary = f"{summary_es} / {summary_en}"
            summary = f"[{sheet_name}] {summary}"

            rows.append(
                {
                    "code": code,
                    "nat": nat,
                    "unit": unit,
                    "summary": summary,
                    "quantity": qty,
                    "price": price,
                    "amount": amount,
                }
            )
    return rows

from __future__ import annotations

import ast
import csv
import json
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output" / "repo_audit"
OUT.mkdir(parents=True, exist_ok=True)


def scan_python_metadata() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for p in ROOT.rglob("*.py"):
        if any(part in {".git", ".venv", "__pycache__"} for part in p.parts):
            continue
        rel = p.relative_to(ROOT).as_posix()
        text = p.read_text(encoding="utf-8", errors="replace")
        try:
            tree = ast.parse(text)
        except SyntaxError:
            continue

        mains: list[int] = []
        funcs: list[dict[str, Any]] = []
        classes: list[dict[str, Any]] = []
        imports: set[str] = set()

        for n in ast.walk(tree):
            if isinstance(n, ast.If):
                try:
                    src = ast.unparse(n.test)
                except Exception:
                    src = ""
                if "__name__" in src and "__main__" in src:
                    mains.append(n.lineno)
            elif isinstance(n, ast.FunctionDef):
                args = [a.arg for a in n.args.args]
                if n.args.vararg:
                    args.append(f"*{n.args.vararg.arg}")
                if n.args.kwarg:
                    args.append(f"**{n.args.kwarg.arg}")
                funcs.append({"name": n.name, "line": n.lineno, "args": args})
            elif isinstance(n, ast.ClassDef):
                classes.append({"name": n.name, "line": n.lineno})
            elif isinstance(n, ast.Import):
                for a in n.names:
                    imports.add(a.name)
            elif isinstance(n, ast.ImportFrom):
                imports.add(n.module or "")

        envs = sorted(set(re.findall(r"os\.getenv\([\"']([^\"']+)", text)))
        urls = sorted(set(re.findall(r"https?://[^\"'\s]+", text)))

        rows.append(
            {
                "file": rel,
                "mains": sorted(set(mains)),
                "functions": sorted(funcs, key=lambda x: x["line"]),
                "classes": sorted(classes, key=lambda x: x["line"]),
                "imports": sorted(imports),
                "env": envs,
                "urls": urls,
            }
        )
    return sorted(rows, key=lambda r: r["file"])


def pres_xlsx_summary() -> dict[str, Any]:
    p = ROOT / "data" / "PRES.xlsx"
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [ws.cell(row=3, column=i).value for i in range(1, ws.max_column + 1)]
    sample = []
    for r in range(4, min(ws.max_row, 12) + 1):
        sample.append([ws.cell(row=r, column=i).value for i in range(1, ws.max_column + 1)])
    return {
        "path": "data/PRES.xlsx",
        "sheet": ws.title,
        "rows": ws.max_row,
        "cols": ws.max_column,
        "header": header,
        "sample_rows": sample,
    }


def construcosto_summary() -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for f in sorted((ROOT / "data" / "construcosto").glob("*.csv")):
        with f.open("r", encoding="utf-8-sig", errors="replace", newline="") as h:
            rd = list(csv.reader(h))
        header = rd[0] if rd else []
        out.append(
            {
                "path": f.relative_to(ROOT).as_posix(),
                "size_bytes": f.stat().st_size,
                "rows": max(len(rd) - 1, 0),
                "cols": len(header),
                "header": header,
                "sample_rows": rd[1:4],
            }
        )
    return out


def data_files_summary() -> list[dict[str, Any]]:
    exts = {".bc3", ".xlsx", ".csv", ".json", ".jsonl", ".rtf", ".jpg", ".png", ".md"}
    out: list[dict[str, Any]] = []
    for p in sorted((ROOT / "data").rglob("*")):
        if not p.is_file():
            continue
        if p.suffix.lower() in exts:
            out.append(
                {
                    "path": p.relative_to(ROOT).as_posix(),
                    "size_bytes": p.stat().st_size,
                    "ext": p.suffix.lower(),
                }
            )
    return out


def main() -> None:
    (OUT / "python_metadata.json").write_text(
        json.dumps(scan_python_metadata(), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT / "pres_xlsx_summary.json").write_text(
        json.dumps(pres_xlsx_summary(), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT / "construcosto_summary.json").write_text(
        json.dumps(construcosto_summary(), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT / "data_files_summary.json").write_text(
        json.dumps(data_files_summary(), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(str(OUT))


if __name__ == "__main__":
    main()

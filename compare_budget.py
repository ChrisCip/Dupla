"""
Compare generated Dupla budget workbook against real PRES.xlsx baseline.

Usage:
    python compare_budget.py
    python compare_budget.py --generated "<path>" --real "./data/NASAS09_Preliminary_Budget.xlsx" --output-dir "<dir>" --real-format nasas_preliminary
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from budget.discipline_mapping import DISCIPLINE_KEYS, SuccessMetricThresholds, canonical_discipline_for_summary
from budget.nasas_preliminary_io import load_nasas_preliminary_budget_rows


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        cleaned = str(value).replace(",", ".").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0


def _normalize(text: str) -> str:
    lowered = text.lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
        "�": "a",
    }
    for src, dst in replacements.items():
        lowered = lowered.replace(src, dst)
    return lowered


def _load_budget_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        code = _safe_str(row[0] if len(row) > 0 else None)
        nat = _safe_str(row[1] if len(row) > 1 else None)
        unit = _safe_str(row[2] if len(row) > 2 else None)
        summary = _safe_str(row[3] if len(row) > 3 else None)
        qty = _safe_float(row[4] if len(row) > 4 else None)
        price = _safe_float(row[5] if len(row) > 5 else None)
        amount = _safe_float(row[6] if len(row) > 6 else None)
        if not any((code, nat, unit, summary)):
            continue
        rows.append(
            {
                "code": code,
                "nat": nat,
                "unit": unit,
                "summary": summary,
                "quantity": qty,
                "price": price,
                "amount": amount,
            }
        )
    return rows


def _is_partida(row: dict[str, Any]) -> bool:
    return "partida" in _normalize(str(row.get("nat", "")))


def _is_chapter(row: dict[str, Any]) -> bool:
    return "cap" in _normalize(str(row.get("nat", "")))


def _discipline_tags(text: str) -> set[str]:
    normalized = _normalize(text)
    tags: dict[str, tuple[str, ...]] = {
        "preliminares": ("prelimin",),
        "movimiento_tierra": ("movimiento de tierra", "excav", "relleno"),
        "hormigon_armado": ("hormigon", "hormigon armado", "concreto"),
        "acero_refuerzo": ("acero", "refuerzo", "varilla"),
        "muros_divisiones": ("muro", "bloque", "division"),
        "panete_revestimiento": ("panete", "pañete", "fraguache", "revest"),
        "pisos": ("piso", "porcelanato", "ceram", "zocalo"),
        "escaleras": ("escalera", "escalon"),
        "puertas": ("puerta",),
        "ventanas": ("ventana", "vidrio"),
        "ebanisteria": ("ebanister", "closet", "gabinete"),
        "electrico": ("electrico", "eléctrico", "luminaria", "tomacorr", "panel"),
        "sanitario": ("sanitario", "inodoro", "lavamanos", "plomer", "drenaje"),
        "pintura": ("pintura", "sellador", "imperme"),
        "techos_cubierta": ("techo", "cubierta"),
        "miscelaneos": ("miscelaneo", "miscelaneo"),
        "equipos_electricos": ("equipos electric",),
        "herreria": ("verja", "baranda", "herreria"),
        "impermeabilizacion": ("impermeabil",),
        "acabados": ("terminacion", "acabado"),
        "gastos_generales": ("gastos", "indirectos", "supervision"),
    }
    found: set[str] = set()
    for tag, hints in tags.items():
        if any(hint in normalized for hint in hints):
            found.add(tag)
    return found


def _line_precision(real_value: float, gen_value: float) -> float:
    if real_value == 0 and gen_value == 0:
        return 1.0
    if real_value == 0:
        return 0.0
    score = 1.0 - abs(gen_value - real_value) / abs(real_value)
    return max(0.0, min(1.0, score))


def _mean(values: list[float]) -> float:
    if not values:
        return 0.0
    return sum(values) / len(values)


def _md_cell(value: Any) -> str:
    text = _safe_str(value).replace("|", "\\|").replace("\n", " ")
    return text


def analyze_budget_pair(
    generated_path: Path,
    real_path: Path,
    *,
    real_format: str = "default",
) -> dict[str, Any]:
    """
    Métricas compartidas entre el informe .txt y el informe Markdown.

    real_format:
      - ``default``: mismo layout Presto que la primera hoja Dupla (filas desde la 4).
      - ``nasas_preliminary``: presupuesto NASAS «Preliminary Budget» (varias hojas).
    """
    generated_rows = _load_budget_rows(generated_path)
    if real_format == "nasas_preliminary":
        real_rows = load_nasas_preliminary_budget_rows(real_path)
    else:
        real_rows = _load_budget_rows(real_path)

    generated_partidas = [row for row in generated_rows if _is_partida(row)]
    real_partidas = [row for row in real_rows if _is_partida(row)]
    generated_chapters = [row for row in generated_rows if _is_chapter(row)]
    real_chapters = [row for row in real_rows if _is_chapter(row)]

    generated_by_code: dict[str, dict[str, Any]] = {
        _safe_str(row["code"]): row for row in generated_partidas if _safe_str(row["code"])
    }
    real_by_code: dict[str, dict[str, Any]] = {
        _safe_str(row["code"]): row for row in real_partidas if _safe_str(row["code"])
    }
    matching_codes = sorted(set(real_by_code) & set(generated_by_code))
    real_only_codes = sorted(set(real_by_code) - set(generated_by_code))
    generated_only_codes = sorted(set(generated_by_code) - set(real_by_code))

    qty_precisions = [
        _line_precision(real_by_code[code]["quantity"], generated_by_code[code]["quantity"])
        for code in matching_codes
    ]
    price_precisions = [
        _line_precision(real_by_code[code]["price"], generated_by_code[code]["price"])
        for code in matching_codes
    ]

    generated_disciplines: set[str] = set()
    for row in generated_rows:
        generated_disciplines.update(_discipline_tags(row["summary"]))
    real_disciplines: set[str] = set()
    for row in real_rows:
        real_disciplines.update(_discipline_tags(row["summary"]))

    top20_real = sorted(real_partidas, key=lambda row: row["amount"], reverse=True)[:20]

    generated_non_empty_price = sum(1 for row in generated_partidas if row["price"] > 0)
    generated_non_empty_amount = sum(1 for row in generated_partidas if row["amount"] > 0)
    real_total = sum(row["amount"] for row in real_partidas)
    generated_total = sum(row["amount"] for row in generated_partidas)

    coverage = 100.0 * (len(matching_codes) / len(real_by_code)) if real_by_code else 0.0
    qty_accuracy = 100.0 * _mean(qty_precisions)
    price_accuracy = 100.0 * _mean(price_precisions)

    amount_deltas: list[dict[str, Any]] = []
    for code in matching_codes:
        r = real_by_code[code]
        g = generated_by_code[code]
        amount_deltas.append(
            {
                "code": code,
                "real_amount": r["amount"],
                "gen_amount": g["amount"],
                "delta": g["amount"] - r["amount"],
                "real_qty": r["quantity"],
                "gen_qty": g["quantity"],
                "summary": _safe_str(r["summary"])[:120],
            }
        )
    amount_deltas.sort(key=lambda item: abs(float(item["delta"])), reverse=True)

    return {
        "generated_rows": generated_rows,
        "real_rows": real_rows,
        "generated_partidas": generated_partidas,
        "real_partidas": real_partidas,
        "generated_chapters": generated_chapters,
        "real_chapters": real_chapters,
        "generated_by_code": generated_by_code,
        "real_by_code": real_by_code,
        "matching_codes": matching_codes,
        "real_only_codes": real_only_codes,
        "generated_only_codes": generated_only_codes,
        "qty_precisions": qty_precisions,
        "price_precisions": price_precisions,
        "generated_disciplines": generated_disciplines,
        "real_disciplines": real_disciplines,
        "top20_real": top20_real,
        "generated_non_empty_price": generated_non_empty_price,
        "generated_non_empty_amount": generated_non_empty_amount,
        "real_total": real_total,
        "generated_total": generated_total,
        "coverage": coverage,
        "qty_accuracy": qty_accuracy,
        "price_accuracy": price_accuracy,
        "amount_deltas": amount_deltas,
    }


def analyze_budget_by_canonical_discipline(
    generated_path: Path,
    real_path: Path,
    *,
    real_format: str = "default",
) -> dict[str, Any]:
    """
    Suma de importes por disciplina canónica (heurística de texto en el resumen).
    Complementa `analyze_budget_pair` para localizar dónde se desvía el total.
    """
    stats = analyze_budget_pair(generated_path, real_path, real_format=real_format)
    buckets: dict[str, dict[str, float]] = {
        k: {"real_amount": 0.0, "generated_amount": 0.0, "real_lines": 0, "generated_lines": 0}
        for k in DISCIPLINE_KEYS
    }
    for row in stats["real_partidas"]:
        d = canonical_discipline_for_summary(str(row.get("summary", "")))
        buckets[d]["real_amount"] += float(row.get("amount", 0) or 0)
        buckets[d]["real_lines"] += 1
    for row in stats["generated_partidas"]:
        d = canonical_discipline_for_summary(str(row.get("summary", "")))
        buckets[d]["generated_amount"] += float(row.get("amount", 0) or 0)
        buckets[d]["generated_lines"] += 1

    th = SuccessMetricThresholds()
    rel_err = (
        abs(stats["generated_total"] - stats["real_total"]) / stats["real_total"]
        if stats["real_total"]
        else 0.0
    )
    return {
        "headline": {
            "coverage_pct": stats["coverage"],
            "qty_accuracy_pct": stats["qty_accuracy"],
            "price_accuracy_pct": stats["price_accuracy"],
            "real_total": stats["real_total"],
            "generated_total": stats["generated_total"],
            "total_amount_rel_error": rel_err,
        },
        "thresholds": {
            "min_code_coverage_pct": th.min_code_coverage_pct,
            "min_qty_accuracy_pct": th.min_qty_accuracy_pct,
            "max_total_amount_rel_error": th.max_total_amount_rel_error,
        },
        "by_discipline": buckets,
    }


def build_comparison_markdown(
    generated_path: Path,
    real_path: Path,
    *,
    title: str,
    run_date: str,
    run_tag: str,
    notes: str = "",
    max_list_codes: int = 80,
    max_delta_rows: int = 25,
    real_format: str = "default",
) -> str:
    """
    Informe en Markdown para carpetas de comparación por proyecto/corrida.
    """
    stats = analyze_budget_pair(generated_path, real_path, real_format=real_format)
    lines: list[str] = [
        f"# {title}",
        "",
        f"- **Fecha de corrida:** {run_date}",
        f"- **Etiqueta de corrida:** `{run_tag}`",
        f"- **Generado (Dupla):** `{generated_path}`",
        f"- **Referencia (PRES):** `{real_path}`",
        "",
        "## Contexto y limitaciones",
        "",
        notes.strip() or (
            "- Esta comparación asume el mismo layout Presto en la primera hoja (filas desde la 4). "
            "- Requiere validación manual si el PRES usa otra hoja o formato."
        ),
        "",
        "## Resumen ejecutivo",
        "",
        "| Métrica | Generado | PRES (real) |",
        "| --- | ---: | ---: |",
        f"| Partidas | {len(stats['generated_partidas'])} | {len(stats['real_partidas'])} |",
        f"| Capítulos (filas Nat) | {len(stats['generated_chapters'])} | {len(stats['real_chapters'])} |",
        f"| Códigos coincidentes | {len(stats['matching_codes'])} | — |",
        f"| Cobertura códigos PRES con equivalente generado | {stats['coverage']:.2f}% | — |",
        f"| Precisión cantidad (solo códigos coincidentes) | {stats['qty_accuracy']:.2f}% | — |",
        f"| Precisión precio unitario (solo coincidentes) | {stats['price_accuracy']:.2f}% | — |",
        f"| Suma Importe (ImpPres) | {stats['generated_total']:,.2f} | {stats['real_total']:,.2f} |",
        f"| Delta (generado − real) | {stats['generated_total'] - stats['real_total']:,.2f} | — |",
        "",
        "### Completitud de precios en generado",
        "",
        f"- Filas partida con PrPres > 0: **{stats['generated_non_empty_price']}** / {len(stats['generated_partidas'])}",
        f"- Filas partida con ImpPres > 0: **{stats['generated_non_empty_amount']}** / {len(stats['generated_partidas'])}",
        "",
        "## Disciplinas (heurística por texto del resumen)",
        "",
        f"- Etiquetas presentes en PRES y no detectadas en generado: **{', '.join(sorted(stats['real_disciplines'] - stats['generated_disciplines'])) or '—'}**",
        "",
        "## Mayores diferencias de importe (códigos en ambos)",
        "",
        "| Código | ImpPres real | ImpPres gen | Delta | Cant. real | Cant. gen | Resumen (PRES) |",
        "| --- | ---: | ---: | ---: | ---: | ---: | --- |",
    ]
    for row in stats["amount_deltas"][:max_delta_rows]:
        lines.append(
            "| {code} | {ra:,.2f} | {ga:,.2f} | {d:,.2f} | {rq} | {gq} | {s} |".format(
                code=_md_cell(row["code"]),
                ra=row["real_amount"],
                ga=row["gen_amount"],
                d=row["delta"],
                rq=row["real_qty"],
                gq=row["gen_qty"],
                s=_md_cell(row["summary"]),
            )
        )
    lines.extend(
        [
            "",
            "## Códigos solo en PRES (no aparecen en generado)",
            "",
        ]
    )
    roc = stats["real_only_codes"]
    if not roc:
        lines.append("_Ninguno._")
    else:
        shown = roc[:max_list_codes]
        lines.extend(f"- `{c}`" for c in shown)
        if len(roc) > max_list_codes:
            lines.append(f"- … y **{len(roc) - max_list_codes}** más.")
    lines.extend(["", "## Códigos solo en generado (no están en PRES)", ""])
    goc = stats["generated_only_codes"]
    if not goc:
        lines.append("_Ninguno._")
    else:
        shown = goc[:max_list_codes]
        lines.extend(f"- `{c}`" for c in shown)
        if len(goc) > max_list_codes:
            lines.append(f"- … y **{len(goc) - max_list_codes}** más.")

    lines.extend(["", "## Top 20 partidas PRES por importe vs generado", ""])
    for row in stats["top20_real"]:
        code = _safe_str(row["code"])
        gen = stats["generated_by_code"].get(code)
        if gen is None:
            lines.append(
                f"- **{code}**: real **{row['amount']:,.2f}** — generado _no encontrado_ — {_md_cell(row['summary'])}"
            )
            continue
        qty_score = 100.0 * _line_precision(row["quantity"], gen["quantity"])
        price_score = 100.0 * _line_precision(row["price"], gen["price"])
        lines.append(
            f"- **{code}**: real **{row['amount']:,.2f}** | gen **{gen['amount']:,.2f}** | "
            f"precisión cant. {qty_score:.1f}% | precio {price_score:.1f}%"
        )

    return "\n".join(lines) + "\n"


def build_comparison_report(
    generated_path: Path,
    real_path: Path,
    output_dir: Path,
    *,
    real_format: str = "default",
) -> Path:
    stats = analyze_budget_pair(generated_path, real_path, real_format=real_format)
    matching_codes = stats["matching_codes"]
    qty_precisions = stats["qty_precisions"]
    price_precisions = stats["price_precisions"]
    generated_disciplines = stats["generated_disciplines"]
    real_disciplines = stats["real_disciplines"]
    top20_real = stats["top20_real"]
    generated_partidas = stats["generated_partidas"]
    real_partidas = stats["real_partidas"]
    generated_chapters = stats["generated_chapters"]
    real_chapters = stats["real_chapters"]
    generated_by_code = stats["generated_by_code"]
    real_by_code = stats["real_by_code"]
    generated_non_empty_price = stats["generated_non_empty_price"]
    generated_non_empty_amount = stats["generated_non_empty_amount"]
    real_total = stats["real_total"]
    generated_total = stats["generated_total"]
    coverage = stats["coverage"]
    qty_accuracy = stats["qty_accuracy"]
    price_accuracy = stats["price_accuracy"]

    lines = [
        "COMPARISON REPORT - DUPLA VS PRES.xlsx",
        "",
        f"Generated workbook: {generated_path}",
        f"Real workbook: {real_path}",
        "",
        "1) High-level counts",
        f"- Partidas generated: {len(generated_partidas)} | real: {len(real_partidas)} (expected ~1565)",
        f"- Chapters generated: {len(generated_chapters)} | real: {len(real_chapters)} (expected ~296)",
        (
            f"- Disciplines covered: {len(generated_disciplines)} | "
            f"real: {len(real_disciplines)} (expected ~21)"
        ),
        "",
        "2) Price/amount completeness in generated",
        f"- Rows with PrPres > 0: {generated_non_empty_price}/{len(generated_partidas)}",
        f"- Rows with ImpPres > 0: {generated_non_empty_amount}/{len(generated_partidas)}",
        "",
        "3) Matching quality by code",
        f"- Matching codes: {len(matching_codes)}",
        f"- Coverage of real partidas by generated equivalent: {coverage:.2f}%",
        f"- Quantity precision (only matched codes): {qty_accuracy:.2f}%",
        f"- Price precision (only matched codes): {price_accuracy:.2f}%",
        "",
        "4) Totals",
        f"- Generated total (sum ImpPres): {generated_total:,.2f}",
        f"- Real total (sum ImpPres): {real_total:,.2f} (reference target: 4,404,786 USD)",
        f"- Delta generated-real: {generated_total - real_total:,.2f}",
        "",
        "5) Missing disciplines (in generated vs real)",
    ]
    missing_disciplines = sorted(real_disciplines - generated_disciplines)
    if missing_disciplines:
        lines.extend(f"- {discipline}" for discipline in missing_disciplines)
    else:
        lines.append("- None")

    lines.extend(
        [
            "",
            "6) Top 20 real partidas by amount vs generated",
        ]
    )
    for row in top20_real:
        code = row["code"]
        gen = generated_by_code.get(code)
        if gen is None:
            lines.append(
                f"- {code}: real_amount={row['amount']:,.2f} | generated=NOT_FOUND | summary={row['summary']}"
            )
            continue
        qty_score = 100.0 * _line_precision(row["quantity"], gen["quantity"])
        price_score = 100.0 * _line_precision(row["price"], gen["price"])
        lines.append(
            (
                f"- {code}: real_amount={row['amount']:,.2f} | gen_amount={gen['amount']:,.2f} | "
                f"qty_precision={qty_score:.2f}% | price_precision={price_score:.2f}%"
            )
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "comparison_report.txt"
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


def _resolve_defaults() -> tuple[Path, Path, Path]:
    repo_root = Path(__file__).resolve().parent
    run_summary = Path(r"C:\Users\chris\Downloads\archivos dupla\dwg\run_summary.json")
    if run_summary.exists():
        payload = json.loads(run_summary.read_text(encoding="utf-8"))
        generated = Path(payload["budget_excel"])
        output_dir = generated.parent
        real = repo_root / "data" / "NASAS09_Preliminary_Budget.xlsx"
        return generated, real, output_dir
    return (
        repo_root / "output" / "dupla_budget_ready_full.xlsx",
        repo_root / "data" / "NASAS09_Preliminary_Budget.xlsx",
        repo_root / "output",
    )


def _validate_xlsx_cli_path(label: str, path: Path) -> None:
    """Evita placeholders tipo `...` y extensiones que openpyxl no abre."""
    raw = str(path).strip()
    if not raw or raw == "..." or "..." in path.name or path.name in {".", ".."}:
        raise SystemExit(
            f"{label}: ruta inválida ({path}). "
            "No uses `...` como placeholder; pasa una ruta real entre comillas, "
            'p. ej. --generated "C:/salida/dupla_budget_ready_full.xlsx"'
        )
    suf = path.suffix.lower()
    if suf not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise SystemExit(
            f"{label}: se esperaba un Excel .xlsx (openpyxl); recibido: {path} "
            f"(extensión {suf!r}). Los .xls antiguos no están soportados."
        )


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Compare generated budget vs PRES.xlsx")
    defaults_generated, defaults_real, defaults_output = _resolve_defaults()
    parser.add_argument("--generated", default=str(defaults_generated))
    parser.add_argument("--real", default=str(defaults_real))
    parser.add_argument("--output-dir", default=str(defaults_output))
    parser.add_argument(
        "--write-markdown",
        type=str,
        default="",
        help="Ruta opcional para informe Markdown (p. ej. comparisons/budget/.../diferencias.md).",
    )
    parser.add_argument(
        "--write-discipline-json",
        type=str,
        default="",
        help="Ruta opcional para JSON con desglose por disciplina canónica.",
    )
    parser.add_argument(
        "--run-tag",
        type=str,
        default="local",
        help="Etiqueta de corrida (solo para el Markdown).",
    )
    parser.add_argument(
        "--real-format",
        choices=("default", "nasas_preliminary"),
        default="nasas_preliminary",
        help="Layout del Excel de referencia: nasas_preliminary (data/NASAS09_Preliminary_Budget.xlsx) o default (Presto, fila 4).",
    )
    return parser


def main() -> None:
    args = _build_arg_parser().parse_args()
    generated = Path(args.generated).resolve()
    real = Path(args.real).resolve()
    output_dir = Path(args.output_dir).resolve()

    _validate_xlsx_cli_path("--generated", generated)
    _validate_xlsx_cli_path("--real", real)

    if not generated.exists():
        raise FileNotFoundError(f"Generated workbook not found: {generated}")
    if not real.exists():
        raise FileNotFoundError(f"Real workbook not found: {real}")

    report_path = build_comparison_report(
        generated, real, output_dir, real_format=args.real_format
    )
    print(f"Comparison report written to: {report_path}")

    if args.write_discipline_json:
        disc_path = Path(args.write_discipline_json).resolve()
        disc_path.parent.mkdir(parents=True, exist_ok=True)
        payload = analyze_budget_by_canonical_discipline(
            generated, real, real_format=args.real_format
        )
        disc_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Discipline breakdown written to: {disc_path}")

    if args.write_markdown:
        from datetime import date

        md_path = Path(args.write_markdown).resolve()
        md_path.parent.mkdir(parents=True, exist_ok=True)
        title = f"Comparación presupuesto — {generated.stem}"
        md = build_comparison_markdown(
            generated,
            real,
            title=title,
            run_date=str(date.today()),
            run_tag=args.run_tag,
            real_format=args.real_format,
        )
        md_path.write_text(md, encoding="utf-8")
        print(f"Markdown report written to: {md_path}")


if __name__ == "__main__":
    main()
